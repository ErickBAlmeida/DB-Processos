import sqlite3

from openpyxl import load_workbook


class Planilha:

    def __init__(self, banco):
        self.planilha = r"C:\Users\ealmeida\Desktop\base ativa agosto.xlsx"
        self.wb = load_workbook(self.planilha)
        self.sheet = self.wb.active
        self.banco = banco

    def identificar_documento(self, linha) -> str:
        
        raiz = str(self.sheet.cell(row=linha, column=7).value)
        filial = str(self.sheet.cell(row=linha, column=8).value)
        num_verificador = str(self.sheet.cell(row=linha, column=9).value)
        
        if filial != '0':
            registro = raiz.zfill(8) + f"000{filial}" + num_verificador.zfill(2)
            tipo_pessoa = 2
            return registro, tipo_pessoa
        
        else:
            registro = raiz.zfill(9) + num_verificador.zfill(2)
            tipo_pessoa = 1
            return registro, tipo_pessoa
    
    def organizar_faixa_de_atraso(self, faixa_de_atraso):

        match faixa_de_atraso:
            case "01.0001-0060":
                return 0
                
            case "01.0001-0080" | "02.0061-0360" | "02.0081-0180" | "03.0181-0360" :
                return 1

            case "03.0361-0720" :
                return 2

            case "04.0721-1080" :
                return 3

            case "05.03 a 05 ANOS" :
                return 4
                
            case "06. > 05 ANOS" | "06.05 a 10 ANOS" :
                return 5

            case "07. > 10 ANOS":
                return 6

            case _ :
                return 

    def extrair_dados(self, linha) -> dict:

            registro, tipo_pessoa = self.identificar_documento(linha)
            nome = self.sheet.cell(row=linha, column=10).value

            devedores = {
                "registro": registro,
                "nome": nome,
                "tp_pessoa": tipo_pessoa
            }

            ag_digital = self.sheet.cell(row=linha, column=22).value
            agencia = self.sheet.cell(row=linha, column=3).value
            conta = self.sheet.cell(row=linha, column=4).value
            carteira = self.sheet.cell(row=linha, column=5).value
            id_gerencia = self.sheet.cell(row=linha, column=23).value

            dados_bancarios = {
                "registro": registro,
                "tp_pessoa": tipo_pessoa,
                "ag_digital" : ag_digital,
                "agencia" : agencia,
                "conta" : conta,
                "carteira" : carteira,
                "id_gerencia" : id_gerencia,
            }

            gcpj = self.sheet.cell(row=linha, column=28).value

            estoque = self.sheet.cell(row=linha, column=13).value
            meta = self.sheet.cell(row=linha, column=11).value
            pago = self.sheet.cell(row=linha, column=12).value
            data_pgto = self.sheet.cell(row=linha, column=20).value
            tipo_pgto = self.sheet.cell(row=linha, column=21).value
            
            atraso = self.sheet.cell(row=linha, column=17).value
            atraso = self.organizar_faixa_de_atraso(atraso)
            
            class_grupo = self.sheet.cell(row=linha, column=14).value
            grupo = self.sheet.cell(row=linha, column=15).value
            segmento = self.sheet.cell(row=linha, column=18).value
            status = self.sheet.cell(row=linha, column=16).value
            contrato = self.sheet.cell(row=linha, column=6).value
            modalidade = self.sheet.cell(row=linha, column=27).value

            processos = {
                "GCPJ" : gcpj,
                "registro": registro,
                "estoque" : estoque,
                "meta" : meta,
                "pago" : pago,
                "data_pgto" : data_pgto,
                "tipo_pgto" : tipo_pgto,
                "atraso" : atraso,
                "class_grupo" : class_grupo,
                "grupo" : grupo,
                "segmento" : segmento,
                "status" : status,
                "contrato" : contrato,
                "modalidade" : modalidade,
            }

            return devedores, dados_bancarios, processos

    def iterar_planilha(self):
        for linha in range(2, self.sheet.max_row+1):        
            if self.sheet.cell(row=linha, column=1).value is None:
                break
                        
            else:                          
                devedores, dados_bancarios, processos = self.extrair_dados(linha)

                self.banco.inserir_dados("devedores", devedores)
                self.banco.inserir_dados("dados_bancarios", dados_bancarios)
                self.banco.inserir_dados("processos", processos)
                print()

class Banco:
    def __init__(self):
        self.banco = sqlite3.connect(r"C:\Users\ealmeida\Desktop\BarrosBD.db")
        self.cursor = self.banco.cursor()
    
    def inserir_dados(self, tabela, dados: dict):
        colunas = ", ".join(dados.keys())
        placeholders = ", ".join(["?" for _ in dados])
        valores = tuple(dados.values())

        sql = f"INSERT INTO {tabela} ({colunas}) VALUES ({placeholders})"

        try:
            self.cursor.execute(sql, valores)
            self.banco.commit()

            print(f"✅ Dados inseridos na tabela '{tabela}' com sucesso!")

        except sqlite3.IntegrityError:
            print(f"❌ O CPF ou CNPJ {dados['registro']} já está cadastrado na tabela '{tabela}'.")

        except sqlite3.Error as e:
            print(f"❌ ERRO GERAL DO SQLITE: {e}")

        except Exception as e:
            print(f"❌ Detalhes do erro: {e}")

    def fechar_banco(self):
        self.cursor.close()
        self.banco.close()

banco = Banco()
planilha = Planilha(banco)
planilha.iterar_planilha()
banco.fechar_banco()